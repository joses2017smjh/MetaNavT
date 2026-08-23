"""Spreadsheet → aggregation plan → user checkpoint → MATLAB chart.

This exposes an auditable decision trace, not private chain-of-thought.
The trace records observable schema facts, the chosen aggregation, chart
recommendation, alternatives, and the approval/override checkpoint.

Generated scripts are MATLAB-compatible. Octave is the portable local runner;
MATLAB is used when explicitly requested and available.
"""

from __future__ import annotations

import csv
import re
import shutil
import subprocess
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from statistics import mean
from typing import Any, Sequence

SUPPORTED_CHARTS = ("bar", "line", "dot", "histogram")
SUPPORTED_OPERATIONS = ("mean", "sum", "min", "max", "count")


@dataclass
class ColumnProfile:
    name: str
    kind: str
    missing: int
    unique: int
    sample: list[Any] = field(default_factory=list)
    minimum: float | None = None
    maximum: float | None = None

    def as_dict(self) -> dict:
        return {
            "name": self.name,
            "kind": self.kind,
            "missing": self.missing,
            "unique": self.unique,
            "sample": self.sample,
            "minimum": self.minimum,
            "maximum": self.maximum,
        }


@dataclass
class VisualizationPlan:
    plan_id: str
    source_path: str
    question: str
    columns: list[ColumnProfile]
    group_by: str
    value: str
    operation: str
    recommended_chart: str
    alternatives: list[str]
    aggregated_rows: list[dict[str, Any]]
    decision_trace: list[dict[str, str]]
    user_questions: list[str]
    matlab_code: str
    script_path: str
    chart_path: str
    baseline: float | None = None
    status: str = "pending_approval"
    approved: bool = False
    applied: bool = False

    def as_dict(self) -> dict:
        return {
            "plan_id": self.plan_id,
            "source_path": self.source_path,
            "question": self.question,
            "columns": [column.as_dict() for column in self.columns],
            "aggregation": {
                "group_by": self.group_by,
                "value": self.value,
                "operation": self.operation,
            },
            "recommended_chart": self.recommended_chart,
            "alternatives": self.alternatives,
            "aggregated_rows": self.aggregated_rows,
            "decision_trace": self.decision_trace,
            "user_questions": self.user_questions,
            "requires_user_input": True,
            "matlab_code": self.matlab_code,
            "script_path": self.script_path,
            "chart_path": self.chart_path,
            "baseline": self.baseline,
            "status": self.status,
        }


def _normalized(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (text or "").lower()).strip()


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _read_spreadsheet(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            headers = list(reader.fieldnames or [])
            return headers, [dict(row) for row in reader]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX support requires openpyxl; CSV/TSV needs no extra dependency") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        return headers, [dict(zip(headers, values)) for values in rows]
    raise ValueError(f"unsupported spreadsheet type: {suffix}; use CSV, TSV, XLSX, or XLSM")


def inspect_spreadsheet(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    headers, rows = _read_spreadsheet(path)
    profiles: list[ColumnProfile] = []
    for header in headers:
        values = [row.get(header) for row in rows]
        present = [value for value in values if value not in (None, "")]
        numeric = [_number(value) for value in present]
        bool_values = {
            str(value).strip().lower()
            for value in present
            if str(value).strip().lower() in {"true", "false"}
        }
        if present and len(bool_values) == len({str(value).strip().lower() for value in present}):
            kind = "boolean"
        elif present and all(value is not None for value in numeric):
            kind = "numeric"
        else:
            kind = "categorical" if len(set(map(str, present))) <= max(20, len(present) // 2) else "text"
        numeric_present = [value for value in numeric if value is not None]
        profiles.append(
            ColumnProfile(
                name=header,
                kind=kind,
                missing=len(values) - len(present),
                unique=len(set(map(str, present))),
                sample=list(dict.fromkeys(present))[:4],
                minimum=min(numeric_present) if numeric_present else None,
                maximum=max(numeric_present) if numeric_present else None,
            )
        )
    return {
        "path": str(path),
        "n_rows": len(rows),
        "n_columns": len(headers),
        "headers": headers,
        "columns": [profile.as_dict() for profile in profiles],
        "rows": rows,
    }


def _mentioned_column(question: str, columns: Sequence[ColumnProfile]) -> str | None:
    query = _normalized(question)
    ranked = []
    for column in columns:
        tokens = _normalized(column.name)
        score = sum(token in query.split() for token in tokens.split())
        if tokens and tokens in query:
            score += 3
        ranked.append((score, column.name))
    ranked.sort(reverse=True)
    return ranked[0][1] if ranked and ranked[0][0] > 0 else None


def _choose_fields(
    question: str,
    columns: Sequence[ColumnProfile],
    *,
    group_by: str | None,
    value: str | None,
) -> tuple[str, str]:
    by_name = {column.name: column for column in columns}
    if group_by and group_by not in by_name:
        raise ValueError(f"unknown group_by column {group_by!r}")
    if value and value not in by_name:
        raise ValueError(f"unknown value column {value!r}")

    categorical = [column for column in columns if column.kind in {"categorical", "boolean", "text"}]
    numeric = [
        column
        for column in columns
        if column.kind == "numeric" and column.name.lower() not in {"id", "run_id"}
    ]
    mentioned = _mentioned_column(question, columns)
    if group_by is None:
        if mentioned and by_name[mentioned].kind != "numeric":
            group_by = mentioned
        else:
            group_by = next((c.name for c in categorical if _normalized(c.name) in _normalized(question)), None)
        group_by = group_by or (categorical[0].name if categorical else columns[0].name)
    if value is None:
        candidates = [c for c in numeric if any(token in _normalized(question) for token in _normalized(c.name).split())]
        value = (candidates[0].name if candidates else numeric[0].name if numeric else "")
    if not value:
        raise ValueError("no numeric metric found; choose a numeric value column")
    if by_name[value].kind != "numeric":
        raise ValueError(f"value column {value!r} is not numeric")
    return group_by, value


def _operation(question: str, explicit: str | None) -> str:
    if explicit:
        if explicit not in SUPPORTED_OPERATIONS:
            raise ValueError(f"unsupported operation {explicit!r}")
        return explicit
    query = _normalized(question)
    if re.search(r"\b(how many|count)\b", query):
        return "count"
    if re.search(r"\b(total|sum)\b", query):
        return "sum"
    if re.search(r"\b(min|minimum|lowest|best)\b", query):
        return "min"
    if re.search(r"\b(max|maximum|highest|worst)\b", query):
        return "max"
    return "mean"


def _aggregate(
    rows: Sequence[dict[str, Any]],
    group_by: str,
    value: str,
    operation: str,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[float]] = {}
    for row in rows:
        group = str(row.get(group_by, ""))
        number = _number(row.get(value))
        if operation != "count" and number is None:
            continue
        grouped.setdefault(group, [])
        if number is not None:
            grouped[group].append(number)
    result = []
    for group in sorted(grouped):
        values = grouped[group]
        if operation == "count":
            aggregated = len(values)
        elif operation == "sum":
            aggregated = sum(values)
        elif operation == "min":
            aggregated = min(values)
        elif operation == "max":
            aggregated = max(values)
        else:
            aggregated = mean(values)
        result.append({group_by: group, f"{operation}_{value}": round(float(aggregated), 6), "n": len(values)})
    return result


def _chart_choice(question: str, explicit: str | None) -> tuple[str, list[str], str]:
    if explicit:
        if explicit not in SUPPORTED_CHARTS:
            raise ValueError(f"unsupported chart type {explicit!r}")
        return explicit, [chart for chart in SUPPORTED_CHARTS if chart != explicit][:2], "user-selected"
    query = _normalized(question)
    if re.search(r"\b(baseline|threshold|target)\b", query):
        return "dot", ["bar", "line"], "point estimates compared with a reference threshold"
    if re.search(r"\b(distribution|histogram|frequency)\b", query):
        return "histogram", ["bar", "dot"], "distribution requested"
    if re.search(r"\b(trend|over time|sequence)\b", query):
        return "line", ["dot", "bar"], "ordered trend requested"
    return "bar", ["dot", "line"], "discrete groups compared on one aggregate metric"


def _baseline(question: str) -> float | None:
    match = re.search(r"\b([01]\.\d+)\s*(?:baseline|threshold|target)\b", question or "", re.I)
    return float(match.group(1)) if match else None


def _matlab_escape(text: str) -> str:
    return str(text).replace("'", "''")


def generate_matlab(
    *,
    source_path: str,
    headers: Sequence[str],
    columns: Sequence[ColumnProfile],
    group_by: str,
    value: str,
    operation: str,
    chart_type: str,
    chart_path: str,
    baseline: float | None,
) -> str:
    if chart_type not in SUPPORTED_CHARTS:
        raise ValueError(f"unsupported chart type {chart_type!r}")
    by_name = {column.name: column for column in columns}
    formats = ["%f" if by_name[name].kind == "numeric" else "%s" for name in headers]
    group_index = headers.index(group_by) + 1
    value_index = headers.index(value) + 1
    output_name = f"{operation} {value}"
    if operation == "mean":
        aggregate_line = "values(i) = mean(metric(mask));"
    elif operation == "sum":
        aggregate_line = "values(i) = sum(metric(mask));"
    elif operation == "min":
        aggregate_line = "values(i) = min(metric(mask));"
    elif operation == "max":
        aggregate_line = "values(i) = max(metric(mask));"
    else:
        aggregate_line = "values(i) = sum(mask);"
    if chart_type == "line":
        plot_lines = "plot(1:numel(groups), values, '-o', 'LineWidth', 1.8);"
    elif chart_type == "dot":
        plot_lines = "scatter(1:numel(groups), values, 70, 'filled');"
    elif chart_type == "histogram":
        plot_lines = "hist(metric, 10);"
    else:
        plot_lines = "bar(values, 0.72);"
    baseline_lines = ""
    if baseline is not None and chart_type != "histogram":
        baseline_lines = (
            f"hold on;\n"
            f"plot([0.5, numel(groups)+0.5], [{baseline}, {baseline}], '--r', 'LineWidth', 1.5);\n"
            f"legend('aggregate', 'baseline {baseline:g}', 'Location', 'best');\n"
            "hold off;\n"
        )
    return f"""% Generated by MetaNaviT after an approval-gated visualization plan.
% Source: {source_path}
script_dir = fileparts(mfilename('fullpath'));
repo_root = fileparts(fileparts(script_dir));
source_path = fullfile(repo_root, '{_matlab_escape(source_path)}');
output_path = fullfile(repo_root, '{_matlab_escape(chart_path)}');
warning('off', 'all');
fid = fopen(source_path, 'r');
assert(fid >= 0, 'Could not open spreadsheet');
data = textscan(fid, '{''.join(formats)}', 'Delimiter', ',', 'HeaderLines', 1);
fclose(fid);
group = data{{{group_index}}};
metric = data{{{value_index}}};
groups = unique(group, 'stable');
values = zeros(numel(groups), 1);
counts = zeros(numel(groups), 1);
for i = 1:numel(groups)
    mask = strcmp(group, groups{{i}});
    counts(i) = sum(mask);
    {aggregate_line}
end

figure('Visible', 'off', 'Color', 'w');
{plot_lines}
if ~strcmp('{chart_type}', 'histogram')
    set(gca, 'XTick', 1:numel(groups), 'XTickLabel', groups);
end
xlabel('{_matlab_escape(group_by)}', 'Interpreter', 'none');
ylabel('{_matlab_escape(output_name)}', 'Interpreter', 'none');
title('{_matlab_escape(output_name)} by {_matlab_escape(group_by)}', 'Interpreter', 'none');
grid on;
{baseline_lines}print(gcf, output_path, '-dpng', '-r140');
close(gcf);
"""


def propose_visualization(
    root: str | Path,
    source_path: str,
    question: str,
    *,
    group_by: str | None = None,
    value: str | None = None,
    operation: str | None = None,
    chart_type: str | None = None,
) -> VisualizationPlan:
    root = Path(root).resolve()
    source = (root / source_path).resolve()
    source.relative_to(root)
    inspected = inspect_spreadsheet(source)
    columns = [ColumnProfile(**column) for column in inspected["columns"]]
    selected_group, selected_value = _choose_fields(
        question,
        columns,
        group_by=group_by,
        value=value,
    )
    selected_operation = _operation(question, operation)
    recommended, alternatives, rationale = _chart_choice(question, chart_type)
    baseline = _baseline(question)
    aggregated = _aggregate(
        inspected["rows"],
        selected_group,
        selected_value,
        selected_operation,
    )
    slug = re.sub(
        r"[^a-z0-9]+",
        "-",
        f"{selected_operation}-{selected_value}-by-{selected_group}".lower(),
    ).strip("-")
    script_path = f"artifacts/visualizations/{slug.replace('-', '_')}.m"
    chart_path = f"artifacts/visualizations/{slug}.png"
    matlab = generate_matlab(
        source_path=source_path,
        headers=inspected["headers"],
        columns=columns,
        group_by=selected_group,
        value=selected_value,
        operation=selected_operation,
        chart_type=recommended,
        chart_path=chart_path,
        baseline=baseline,
    )
    trace = [
        {
            "stage": "inspect",
            "evidence": f"{inspected['n_rows']} rows; {inspected['n_columns']} columns",
            "decision": "profile numeric and categorical fields before aggregation",
        },
        {
            "stage": "aggregate",
            "evidence": f"group={selected_group}; metric={selected_value}",
            "decision": f"{selected_operation}({selected_value}) by {selected_group}",
        },
        {
            "stage": "visualize",
            "evidence": rationale,
            "decision": f"recommend {recommended}; alternatives: {', '.join(alternatives)}",
        },
        {
            "stage": "checkpoint",
            "evidence": "no script or chart has been written",
            "decision": "wait for approval or a chart override",
        },
    ]
    questions = [
        f"Use {selected_operation}({selected_value}) grouped by {selected_group}?",
        f"Use the recommended {recommended} chart, or override with {', '.join(alternatives)}?",
        f"Write {script_path} and render {chart_path}?",
    ]
    return VisualizationPlan(
        plan_id=str(uuid.uuid4()),
        source_path=source_path,
        question=question,
        columns=columns,
        group_by=selected_group,
        value=selected_value,
        operation=selected_operation,
        recommended_chart=recommended,
        alternatives=alternatives,
        aggregated_rows=aggregated,
        decision_trace=trace,
        user_questions=questions,
        matlab_code=matlab,
        script_path=script_path,
        chart_path=chart_path,
        baseline=baseline,
    )


def available_backend(preferred: str = "auto") -> tuple[str | None, str | None]:
    if preferred not in {"auto", "octave", "matlab"}:
        raise ValueError("backend must be auto, octave, or matlab")
    order = ("matlab", "octave") if preferred == "auto" else (preferred,)
    for name in order:
        executable = shutil.which(name)
        if executable:
            return name, executable
    return None, None


def execute_matlab(
    root: str | Path,
    script_path: str,
    *,
    chart_path: str | None = None,
    backend: str = "auto",
    timeout_s: int = 90,
) -> dict[str, Any]:
    root = Path(root).resolve()
    script = (root / script_path).resolve()
    script.relative_to(root)
    chart = (root / chart_path).resolve() if chart_path else None
    if chart is not None:
        chart.relative_to(root)
    name, executable = available_backend(backend)
    if not executable or not name:
        return {
            "ok": False,
            "backend": None,
            "stdout": "",
            "stderr": "MATLAB/Octave not found",
            "returncode": None,
        }
    if name == "octave":
        command = [executable, "--no-gui", "--quiet", str(script)]
    else:
        escaped = str(script).replace("'", "''")
        command = [executable, "-batch", f"run('{escaped}')"]
    try:
        result = subprocess.run(
            command,
            cwd=root,
            text=True,
            capture_output=True,
            timeout=timeout_s,
            check=False,
        )
        return {
            "ok": result.returncode == 0,
            "backend": name,
            "stdout": result.stdout,
            "stderr": result.stderr,
            "returncode": result.returncode,
        }
    except subprocess.TimeoutExpired as exc:
        rendered = bool(chart and chart.is_file() and chart.stat().st_size > 0)
        return {
            "ok": rendered,
            "backend": name,
            "stdout": exc.stdout or "",
            "stderr": (
                f"renderer timed out after {timeout_s}s after writing the chart"
                if rendered
                else f"timed out after {timeout_s}s"
            ),
            "returncode": None,
        }
